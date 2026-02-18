from fastapi import FastAPI, Request, Depends, HTTPException, Form, status
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
import pandas as pd
import numpy as np
from typing import Optional
import os
from supabase import create_client, Client
import secrets
from dotenv import load_dotenv
import io
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

from regulatory_data import (
    CATEGORIAS_REGULATORIAS, 
    PAISES_DISPONIBLES, 
    extraer_info_regulatoria_pdf,
    get_available_countries,
    get_available_categories,
    get_subcategories,
    get_regulatory_info
)

# Cargar variables de entorno desde .env
load_dotenv()

# Configuración de la aplicación
app = FastAPI(title="Portal de Dashboards")

# Middleware para sesiones
app.add_middleware(SessionMiddleware, secret_key=secrets.token_hex(32))

# Configuración de archivos estáticos y templates (crear carpetas si no existen)
if not os.path.exists("static"):
    os.makedirs("static")
if not os.path.exists("templates"):
    os.makedirs("templates")
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# Configuración de Supabase
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

print(f"🔍 Debug - SUPABASE_URL: {SUPABASE_URL}")
print(f"🔍 Debug - SUPABASE_KEY presente: {'Sí' if SUPABASE_KEY else 'No'}")

# Validar formato de URL
if SUPABASE_URL and not SUPABASE_URL.startswith('https://'):
    print("❌ Error: SUPABASE_URL debe comenzar con 'https://'")
    DEVELOPMENT_MODE = True
elif not SUPABASE_URL or not SUPABASE_KEY:
    print("❌ Error: Variables SUPABASE_URL o SUPABASE_KEY no encontradas")
    DEVELOPMENT_MODE = True
else:
    DEVELOPMENT_MODE = False

if not DEVELOPMENT_MODE:
    try:
        supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
        print("✅ Conectado a Supabase exitosamente")
    except Exception as e:
        print(f"❌ Error conectando a Supabase: {e}")
        print("🔄 Cambiando a modo desarrollo...")
        DEVELOPMENT_MODE = True
        supabase = None
else:
    supabase = None
    print("⚠️  Modo desarrollo activado")
    print("🔍 Credenciales de prueba: admin@test.com / password123")

# Cache para datos
@app.on_event("startup")
def load_data_cache():
    global df_cache, df_suplementos_cache, df_referencias_cache, df_otras_sustancias_cache, df_actualizacion_regulatoria_cache
    
    # Cargar datos de moléculas (existente)
    try:
        df_cache = pd.read_excel('moleculas_sin_duplicados_2025.xlsx', 
                               sheet_name='Hoja1')
        df_cache = clean_duplicates(df_cache)
        print(f"Datos moléculas cargados: {len(df_cache)} registros")
    except FileNotFoundError:
        print("⚠️ Archivo Excel no encontrado. Usando datos de ejemplo.")
        df_cache = create_sample_data()
    except Exception as e:
        print(f"Error cargando datos de moléculas: {e}")
        df_cache = create_sample_data()
    
    # Cargar datos de suplementos
    try:
        # Cargar datos principales de suplementos
        df_suplementos_cache = pd.read_csv('suplementos_normalizados_completo.csv')
        print(f"Datos suplementos cargados: {len(df_suplementos_cache)} registros")
        
        # Cargar referencias de vitaminas
        df_referencias_vitaminas = pd.read_csv('referencias_suplementos_vitaminas.csv')
        print(f"Referencias vitaminas cargadas: {len(df_referencias_vitaminas)} registros")
        
        # Cargar referencias de minerales
        df_referencias_minerales = pd.read_csv('referencias_suplementos_minerales.csv')
        print(f"Referencias minerales cargadas: {len(df_referencias_minerales)} registros")
        
        # Combinar referencias
        df_referencias_cache = pd.concat([df_referencias_vitaminas, df_referencias_minerales], 
                                       ignore_index=True)
        print(f"Total referencias: {len(df_referencias_cache)} registros")
        
    except FileNotFoundError as e:
        print(f"⚠️ Archivo de suplementos no encontrado: {e}")
        print("Usando datos de ejemplo...")
        df_suplementos_cache = create_sample_suplementos_data()
        df_referencias_cache = create_sample_referencias_data()
    except Exception as e:
        print(f"Error cargando datos de suplementos: {e}")
        print("Usando datos de ejemplo...")
        df_suplementos_cache = create_sample_suplementos_data()
        df_referencias_cache = create_sample_referencias_data()

    # Cargar datos de otras sustancias
    try:
        df_otras_sustancias_cache = pd.read_excel('otras_substancias_ilar.xlsx', 
                                                 sheet_name='Otras Substancias')
        print(f"Datos otras sustancias cargados: {len(df_otras_sustancias_cache)} registros")
        
        # Verificar estructura
        print(f"Columnas disponibles: {list(df_otras_sustancias_cache.columns)}")
        print(f"Categorías únicas: {df_otras_sustancias_cache['Category'].unique()}")
        
    except FileNotFoundError:
        print("⚠️ Archivo otras_substancias_ilar.xlsx no encontrado. Usando datos de ejemplo.")
        df_otras_sustancias_cache = create_sample_otras_sustancias_data()
    except Exception as e:
        print(f"Error cargando datos de otras sustancias: {e}")
        df_otras_sustancias_cache = create_sample_otras_sustancias_data()
    
    # Cargar datos de actualización regulatoria
    try:
        df_actualizacion_regulatoria_cache = load_actualizacion_regulatoria_from_excel()
        print(f"Datos actualización regulatoria cargados: {len(df_actualizacion_regulatoria_cache)} registros")
        
    except Exception as e:
        print(f"Error cargando datos de actualización regulatoria: {e}")
        print("Usando datos de ejemplo...")
        df_actualizacion_regulatoria_cache = create_sample_actualizacion_regulatoria_data()

def clean_duplicates(df):
    """Limpia duplicados basándose en columnas clave"""
    key_columns = ['Ingrediente activo', 'País', 'Año de comercialización', 'Dosis permitidas/Duración']
    df_cleaned = df.drop_duplicates()
    df_cleaned = df_cleaned.drop_duplicates(subset=key_columns, keep='first')
    return df_cleaned

def create_sample_data():
    """Crea datos de ejemplo para moléculas"""
    import random
    countries = ['Spain', 'France', 'Germany', 'Italy', 'Netherlands', 'Belgium']
    ingredientes = ['Ibuprofen', 'Paracetamol', 'Aspirin', 'Omeprazole', 'Simvastatin']
    vias = ['Oral', 'Tópica', 'Sublingual', 'Transdérmica']
    clasificaciones = ['RX', 'OTC', 'Rx-OTC']
    
    data = []
    for _ in range(100):
        data.append({
            'País': random.choice(countries),
            'Ingrediente activo': random.choice(ingredientes),
            'Vía de administración': random.choice(vias),
            'Dosis permitidas/Duración': f"{random.randint(10, 500)}mg / {random.randint(3, 14)} días",
            'Indicación ingrediente 1': f"Indicación {random.randint(1, 10)}",
            'Indicación ingrediente 2': f"Indicación {random.randint(1, 10)}",
            'Combinaciones registradas con el ingrediente activo': f"Combinación con ingrediente {random.randint(1, 5)}",
            'Indicación producto/Declaración de propiedades': f"Tratamiento de condición {random.randint(1, 15)}",
            'Año de comercialización': random.randint(2010, 2023),
            'Clasificación regulatoria': random.choice(clasificaciones)
        })
    return pd.DataFrame(data)

def create_sample_suplementos_data():
    """Crea datos de ejemplo para suplementos"""
    import random
    paises = ['Argentina', 'Brasil', 'Chile', 'Colombia', 'México', 'Perú']
    ingredientes_vitaminas = ['Vitamina A', 'Vitamina C', 'Vitamina D', 'Vitamina E', 'Vitamina B12']
    ingredientes_minerales = ['Calcio', 'Hierro', 'Zinc', 'Magnesio', 'Selenio']
    categorias = ['Suplemento Dietario', 'Alimento Funcional', 'Medicamento OTC']
    
    data = []
    for pais in paises:
        for ingrediente in ingredientes_vitaminas + ingredientes_minerales:
            tipo = 'Vitaminas' if ingrediente.startswith('Vitamina') else 'Minerales'
            establecido = random.choice([True, False])
            
            if establecido:
                minimo = round(random.uniform(0.1, 10), 3)
                maximo = round(minimo + random.uniform(5, 50), 3)
                referencias = random.randint(1, 20)
                valor_original = f"{minimo}-{maximo}"
            else:
                minimo = 0.0  # Usar 0.0 en lugar de None
                maximo = 0.0  # Usar 0.0 en lugar de None
                referencias = 0  # Usar 0 en lugar de None
                valor_original = ""
            
            data.append({
                'pais': pais,
                'ingrediente': ingrediente,
                'tipo': tipo,
                'unidad': 'mg' if tipo == 'Minerales' else 'μg',
                'minimo': minimo,
                'maximo': maximo,
                'establecido': establecido,
                'categoria_regulacion': random.choice(categorias),
                'referencias': referencias,
                'valor_original': valor_original
            })
    
    return pd.DataFrame(data)

def create_sample_referencias_data():
    """Crea datos de ejemplo para referencias"""
    data = []
    for i in range(1, 21):
        for tipo in ['Vitaminas', 'Minerales']:
            data.append({
                'referencia': i,
                'descripcion': f"Norma regulatoria {i} para {tipo.lower()}",
                'tipo': tipo
            })
    
    return pd.DataFrame(data)


def create_branded_excel(
    df: pd.DataFrame,
    title: str,
    subtitle: str,
    section_name: str,
    column_widths: dict = None
) -> io.BytesIO:
    """Genera un archivo Excel con formato de tabla, encabezado con títulos y logos de ILAR."""
    wb = Workbook()
    ws = wb.active
    ws.title = section_name[:31]

    # Colores de marca
    ILAR_DARK_BLUE = "2D4A65"
    ILAR_GREEN = "8BA873"
    HEADER_BG = PatternFill(start_color=ILAR_DARK_BLUE, end_color=ILAR_DARK_BLUE, fill_type="solid")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(name="Calibri", bold=True, color=ILAR_DARK_BLUE, size=16)
    SUBTITLE_FONT = Font(name="Calibri", color="666666", size=11)
    SECTION_FONT = Font(name="Calibri", bold=True, color=ILAR_GREEN, size=13)
    LINK_FONT = Font(name="Calibri", color=ILAR_GREEN, size=10, underline="single")
    DATE_FONT = Font(name="Calibri", color="999999", size=9)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    num_cols = max(len(df.columns), 4)

    # --- Encabezado (filas 1-7) ---
    # Fila 1: espaciador
    ws.row_dimensions[1].height = 6

    # Fila 2: Título
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(num_cols - 2, 5))
    cell_title = ws.cell(row=2, column=1, value=title)
    cell_title.font = TITLE_FONT
    cell_title.alignment = Alignment(vertical="center")

    # Fila 3: Subtítulo
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=min(num_cols - 2, 5))
    cell_sub = ws.cell(row=3, column=1, value=subtitle)
    cell_sub.font = SUBTITLE_FONT
    cell_sub.alignment = Alignment(vertical="center")

    # Fila 4: Nombre de sección
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=min(num_cols - 2, 5))
    cell_section = ws.cell(row=4, column=1, value=section_name)
    cell_section.font = SECTION_FONT
    cell_section.alignment = Alignment(vertical="center")

    # Fila 5: Link Portal ILAR
    cell_link = ws.cell(row=5, column=1, value="Portal ILAR")
    cell_link.font = LINK_FONT
    cell_link.hyperlink = "https://portalilar.org"

    # Fila 6: Fecha de generación
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
    cell_date = ws.cell(row=6, column=1, value=f"Generado: {fecha}")
    cell_date.font = DATE_FONT

    # Fila 7: espaciador
    ws.row_dimensions[7].height = 10

    # --- Insertar logos (imagen original sin perder resolución) ---
    from PIL import Image as PILImage

    logo_path = os.path.join("static", "images", "logo-ilar.png")
    if os.path.exists(logo_path):
        try:
            with PILImage.open(logo_path) as pil_img:
                orig_w, orig_h = pil_img.size
            logo_img = XlImage(logo_path)
            # Visualizar a 250px ancho manteniendo proporción (imagen original intacta)
            logo_img.width = 250
            logo_img.height = int(orig_h * 250 / orig_w)
            anchor_col = get_column_letter(min(num_cols, 6))
            ws.add_image(logo_img, f"{anchor_col}1")
        except Exception:
            pass

    banner_path = os.path.join("static", "images", "banner-suplementos.png")
    if os.path.exists(banner_path):
        try:
            with PILImage.open(banner_path) as pil_img:
                orig_w, orig_h = pil_img.size
            banner_img = XlImage(banner_path)
            # Visualizar a 200px ancho manteniendo proporción (imagen original intacta)
            banner_img.width = 200
            banner_img.height = int(orig_h * 200 / orig_w)
            anchor_col_banner = get_column_letter(min(num_cols + 1, 8))
            ws.add_image(banner_img, f"{anchor_col_banner}1")
        except Exception:
            pass

    # --- Tabla de datos (fila 8+) ---
    data_start_row = 8

    # Encabezados de columna
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=data_start_row, column=col_idx, value=str(col_name))
        cell.font = HEADER_FONT
        cell.fill = HEADER_BG
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Datos
    alt_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, (_, row_data) in enumerate(df.iterrows(), data_start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = data_align
            if (row_idx - data_start_row) % 2 == 0:
                cell.fill = alt_fill

    # --- Ancho de columnas automático ---
    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(df.columns[col_idx - 1]))

        for row in ws.iter_rows(min_row=data_start_row + 1, max_row=min(ws.max_row, data_start_row + 100),
                                min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 50))

        adjusted_width = min(max_length + 4, 60)

        if column_widths and col_idx in column_widths:
            adjusted_width = column_widths[col_idx]

        ws.column_dimensions[col_letter].width = adjusted_width

    # Frozen panes
    ws.freeze_panes = f"A{data_start_row + 1}"

    # Configuración de página
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # Guardar en BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def get_current_user(request: Request) -> Optional[str]:
    """Obtiene el usuario actual de la sesión"""
    return request.session.get("user_email")

def require_auth(request: Request):
    """Middleware para requerir autenticación"""
    user = get_current_user(request)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="No autenticado"
        )
    return user

def calculate_avg_authorized_ingredients():
    """Calcula el promedio de ingredientes autorizados por país"""
    global df_suplementos_cache
    
    if df_suplementos_cache is None or len(df_suplementos_cache) == 0:
        return 0.0
    
    # Filtrar solo los ingredientes establecidos (autorizados)
    authorized_data = df_suplementos_cache[df_suplementos_cache['establecido'] == 'Sí']
    
    if len(authorized_data) == 0:
        return 0.0
    
    # Contar ingredientes autorizados por país
    ingredients_per_country = authorized_data.groupby('pais')['ingrediente'].count()
    
    # Calcular el promedio
    average_authorized = ingredients_per_country.mean()
    
    return round(average_authorized, 1)

def calculate_avg_authorized_other_substances():
    """Calcula el promedio de otras sustancias autorizadas por país"""
    global df_otras_sustancias_cache
    
    if df_otras_sustancias_cache is None or len(df_otras_sustancias_cache) == 0:
        return 0.0
    
    # Obtener las columnas de países (todas excepto Category y Substance)
    country_columns = [col for col in df_otras_sustancias_cache.columns 
                      if col not in ['Category', 'Substance']]
    
    if len(country_columns) == 0:
        return 0.0
    
    # Contar sustancias autorizadas por país
    authorized_counts = {}
    
    for country in country_columns:
        # Contar cuántas sustancias están "Autorizado" en este país
        authorized_count = df_otras_sustancias_cache[
            df_otras_sustancias_cache[country] == 'Autorizado'
        ].shape[0]
        
        authorized_counts[country] = authorized_count
    
    # Calcular el promedio
    if len(authorized_counts) > 0:
        total_authorized = sum(authorized_counts.values())
        average_authorized = total_authorized / len(authorized_counts)
        return round(average_authorized, 1)
    
    return 0.0

# Rutas públicas
@app.get("/", response_class=HTMLResponse)
async def login_page(request: Request):
    """Página de login"""
    user = get_current_user(request)
    if user:
        return RedirectResponse(url="/dashboard", status_code=302)
    
    return templates.TemplateResponse("login.html", {"request": request})

@app.post("/login")
async def login(request: Request, email: str = Form(...), password: str = Form(...)):
    """Procesar login"""
    try:
        if DEVELOPMENT_MODE:
            # Modo desarrollo: credenciales fijas
            if email == "admin@test.com" and password == "password123":
                request.session["user_email"] = email
                request.session["user_id"] = "dev_user_123"
                return RedirectResponse(url="/dashboard", status_code=302)
            else:
                raise Exception("Credenciales de desarrollo inválidas")
        else:
            # Modo producción: usar Supabase
            response = supabase.auth.sign_in_with_password({
                "email": email,
                "password": password
            })
            
            if response.user:
                request.session["user_email"] = email
                request.session["user_id"] = response.user.id
                return RedirectResponse(url="/dashboard", status_code=302)
            else:
                raise Exception("Credenciales inválidas")
                
    except Exception as e:
        error_msg = "Credenciales inválidas"
        if DEVELOPMENT_MODE:
            error_msg = "Usa: admin@test.com / password123"
            
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": error_msg
        })

@app.get("/logout")
async def logout(request: Request):
    """Cerrar sesión"""
    request.session.clear()
    return RedirectResponse(url="/", status_code=302)

# Rutas protegidas
@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard_home(request: Request, user: str = Depends(require_auth)):
    """Página principal del dashboard"""
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "user": user
    })

@app.get("/dashboard/molecules", response_class=HTMLResponse)
async def molecules_dashboard(request: Request, user: str = Depends(require_auth)):
    """Dashboard de moléculas ILAR"""
    global df_cache
    
    # Obtener listas para filtros
    molecules = sorted(df_cache['Ingrediente activo'].unique().tolist())
    countries = sorted(df_cache['País'].unique().tolist())
    
    return templates.TemplateResponse("molecules_dashboard.html", {
        "request": request,
        "user": user,
        "molecules": molecules,
        "countries": countries
    })

@app.get("/dashboard/suplementos", response_class=HTMLResponse)
async def suplementos_dashboard(request: Request, user: str = Depends(require_auth)):
    """Dashboard de suplementos América Latina"""
    return templates.TemplateResponse("suplementos_dashboard.html", {
        "request": request,
        "user": user
    })

# APIs existentes para moléculas (sin cambios)
@app.get("/api/molecules-data")
async def get_molecules_data(
    request: Request,
    molecule: Optional[str] = None,
    countries: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    user: str = Depends(require_auth)
):
    """API para obtener datos filtrados del dashboard de moléculas con paginación"""
    global df_cache
    
    print(f"🔍 Debug - Filtros recibidos:")
    print(f"   - Molécula: '{molecule}'")
    print(f"   - Países: '{countries}'")
    print(f"   - Página: {page}")
    
    # Aplicar filtros
    filtered_df = df_cache.copy()
    original_count = len(filtered_df)
    
    print(f"   - Total registros iniciales: {original_count}")
    print(f"   - Moléculas únicas disponibles: {sorted(filtered_df['Ingrediente activo'].unique())}")
    
    # Filtrar por molécula
    if molecule and molecule != "all" and molecule.strip() != "":
        print(f"   - Filtrando por molécula exacta: '{molecule}'")
        
        # Verificar si la molécula existe exactamente en los datos
        available_molecules = filtered_df['Ingrediente activo'].unique()
        exact_match = molecule in available_molecules
        
        print(f"   - ¿Molécula '{molecule}' existe en datos?: {exact_match}")
        
        if exact_match:
            filtered_df = filtered_df[filtered_df['Ingrediente activo'] == molecule]
            print(f"   - Registros después de filtrar por molécula: {len(filtered_df)}")
        else:
            print(f"   - ⚠️ Molécula '{molecule}' no encontrada. Moléculas disponibles:")
            for mol in sorted(available_molecules):
                print(f"     - '{mol}'")
            # Si no se encuentra la molécula, devolver DataFrame vacío
            filtered_df = filtered_df[filtered_df['Ingrediente activo'] == 'MOLÉCULA_NO_ENCONTRADA']
    
    # Filtrar por países
    if countries and countries.strip():
        country_list = [c.strip() for c in countries.split(',') if c.strip()]
        if country_list:
            print(f"   - Filtrando por países: {country_list}")
            available_countries = filtered_df['País'].unique()
            valid_countries = [c for c in country_list if c in available_countries]
            print(f"   - Países válidos encontrados: {valid_countries}")
            
            if valid_countries:
                filtered_df = filtered_df[filtered_df['País'].isin(valid_countries)]
                print(f"   - Registros después de filtrar por países: {len(filtered_df)}")
            else:
                print(f"   - ⚠️ Ningún país válido encontrado")
                filtered_df = filtered_df[filtered_df['País'] == 'PAÍS_NO_ENCONTRADO']
    
    # Métricas básicas
    total_records = len(filtered_df)
    unique_countries = filtered_df['País'].nunique() if total_records > 0 else 0
    unique_molecules = filtered_df['Ingrediente activo'].nunique() if total_records > 0 else 0
    
    print(f"   - Total registros finales: {total_records}")
    print(f"   - Países únicos: {unique_countries}")
    print(f"   - Moléculas únicas: {unique_molecules}")
    
    # Paginación
    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    paginated_df = filtered_df.iloc[start_idx:end_idx]
    
    # Datos para la tabla (solo la página actual)
    if total_records > 0:
        # Asegurar que las columnas existen antes de seleccionarlas
        required_columns = [
            'País', 
            'Ingrediente activo', 
            'Vía de administración', 
            'Dosis permitidas/Duración',
            'Indicación ingrediente 1',
            'Indicación ingrediente 2', 
            'Combinaciones registradas con el ingrediente activo',
            'Indicación producto/Declaración de propiedades',
            'Año de comercialización',
            'Clasificación regulatoria'
        ]
        available_columns = [col for col in required_columns if col in paginated_df.columns]
        
        if len(available_columns) != len(required_columns):
            print(f"   - ⚠️ Columnas faltantes: {set(required_columns) - set(available_columns)}")
            print(f"   - Columnas disponibles: {list(paginated_df.columns)}")
        
        table_data = paginated_df[available_columns].fillna('').to_dict('records')
    else:
        table_data = []
    
    # Información de paginación
    total_pages = (total_records + page_size - 1) // page_size if total_records > 0 else 1
    
    # Asegurar que la página actual no exceda el total
    if page > total_pages and total_pages > 0:
        page = total_pages
    
    return {
        "success": True,
        "filters_applied": {
            "molecule": molecule,
            "countries": countries,
            "molecule_found": molecule in df_cache['Ingrediente activo'].unique() if molecule and molecule != "all" else True
        },
        "metrics": {
            "total_records": total_records,
            "unique_countries": unique_countries,
            "unique_molecules": unique_molecules
        },
        "table_data": table_data,
        "pagination": {
            "current_page": page,
            "total_pages": total_pages,
            "page_size": page_size,
            "has_previous": page > 1,
            "has_next": page < total_pages,
            "showing_from": start_idx + 1 if total_records > 0 else 0,
            "showing_to": min(end_idx, total_records),
            "total_records": total_records
        }
    }

# APIs nuevas para suplementos
@app.get("/api/suplementos-initial")
async def get_suplementos_initial(request: Request, user: str = Depends(require_auth)):
    """API para obtener datos iniciales del dashboard de suplementos"""
    try:
        global df_suplementos_cache, df_referencias_cache
        
        # Verificar que las variables globales existan
        if 'df_suplementos_cache' not in globals() or df_suplementos_cache is None:
            print("Creando datos de ejemplo para suplementos...")
            df_suplementos_cache = create_sample_suplementos_data()
        
        if 'df_referencias_cache' not in globals() or df_referencias_cache is None:
            print("Creando datos de ejemplo para referencias...")
            df_referencias_cache = create_sample_referencias_data()
        
        # Limpiar NaN antes de convertir a dict
        df_suplementos_clean = df_suplementos_cache.fillna(0)  # Reemplazar NaN con 0
        df_referencias_clean = df_referencias_cache.fillna("")  # Reemplazar NaN con string vacío
        
        return {
            "success": True,
            "data": df_suplementos_clean.to_dict('records'),
            "references": df_referencias_clean.to_dict('records')
        }
        
    except Exception as e:
        print(f"Error en suplementos-initial: {e}")
        # En caso de cualquier error, devolver datos de ejemplo limpios
        sample_data = create_sample_suplementos_data()
        sample_references = create_sample_referencias_data()
        
        # Limpiar NaN
        sample_data_clean = sample_data.fillna(0)
        sample_references_clean = sample_references.fillna("")
        
        return {
            "success": True,
            "data": sample_data_clean.to_dict('records'),
            "references": sample_references_clean.to_dict('records')
        }

# AGREGAR NUEVA API PARA DATOS INICIALES DE OTRAS SUSTANCIAS
@app.get("/api/otras-sustancias-initial")
async def get_otras_sustancias_initial(request: Request, user: str = Depends(require_auth)):
    """API para obtener datos iniciales de otras sustancias"""
    try:
        global df_otras_sustancias_cache
        
        if 'df_otras_sustancias_cache' not in globals() or df_otras_sustancias_cache is None:
            print("Creando datos de ejemplo para otras sustancias...")
            df_otras_sustancias_cache = create_sample_otras_sustancias_data()
        
        # Obtener listas únicas para filtros
        categories = sorted(df_otras_sustancias_cache['Category'].unique().tolist())
        
        # Los países son todas las columnas excepto Category y Substance
        country_columns = [col for col in df_otras_sustancias_cache.columns 
                          if col not in ['Category', 'Substance']]
        countries = sorted(country_columns)
        
        return {
            "success": True,
            "categories": categories,
            "countries": countries,
            "total_records": len(df_otras_sustancias_cache)
        }
        
    except Exception as e:
        print(f"Error en otras-sustancias-initial: {e}")
        return {
            "success": False,
            "error": str(e),
            "categories": [],
            "countries": [],
            "total_records": 0
        }

@app.get("/api/suplementos-analysis")
async def get_suplementos_analysis(
    request: Request,
    tipo: Optional[str] = None,
    ingredientes: Optional[str] = None,
    paises: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    user: str = Depends(require_auth)
):
    """API para análisis de suplementos con filtros y paginación"""
    try:
        global df_suplementos_cache, df_referencias_cache
        
        if df_suplementos_cache is None:
            df_suplementos_cache = create_sample_suplementos_data()
        
        if df_referencias_cache is None:
            df_referencias_cache = create_sample_referencias_data()
        
        # Aplicar filtros
        filtered_df = df_suplementos_cache.copy()
        
        if tipo and tipo != "all":
            filtered_df = filtered_df[filtered_df['tipo'] == tipo]
        
        if ingredientes and ingredientes.strip():
            ingrediente_list = [i.strip() for i in ingredientes.split(',') if i.strip()]
            if ingrediente_list:
                filtered_df = filtered_df[filtered_df['ingrediente'].isin(ingrediente_list)]
        
        if paises and paises.strip():
            pais_list = [p.strip() for p in paises.split(',') if p.strip()]
            if pais_list:
                filtered_df = filtered_df[filtered_df['pais'].isin(pais_list)]
        
        # NUEVA FUNCIONALIDAD: Hacer JOIN con referencias
        def get_reference_text(ref_num, tipo):
            """Obtiene el texto de la referencia basado en número y tipo"""
            if pd.isna(ref_num) or ref_num == 0:
                return "-"
            
            try:
                ref_num = int(ref_num)
                ref_row = df_referencias_cache[
                    (df_referencias_cache['referencia'] == ref_num) & 
                    (df_referencias_cache['tipo'] == tipo)
                ]
                
                if not ref_row.empty:
                    descripcion = ref_row.iloc[0]['descripcion']
                    return f"{ref_num}: {descripcion}"
                else:
                    return f"{ref_num}: Referencia no encontrada"
            except:
                return str(ref_num)
        
        # Aplicar el JOIN a los datos filtrados
        filtered_df = filtered_df.copy()
        filtered_df['referencias_texto'] = filtered_df.apply(
            lambda row: get_reference_text(row['referencias'], row['tipo']), 
            axis=1
        )
        
        # Paginación
        total_records = len(filtered_df)
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_df = filtered_df.iloc[start_idx:end_idx]
        
        # Limpiar NaN antes de convertir a dict
        paginated_df_clean = paginated_df.fillna(0)
        
        # Seleccionar columnas para la tabla (incluyendo referencias_texto)
        columns_for_table = ['pais', 'ingrediente', 'tipo', 'minimo', 'maximo', 
                           'unidad', 'establecido', 'categoria_regulacion', 'referencias_texto']
        
        table_data = []
        for _, row in paginated_df_clean.iterrows():
            table_data.append({
                'pais': row['pais'],
                'ingrediente': row['ingrediente'],
                'tipo': row['tipo'],
                'minimo': row['minimo'],
                'maximo': row['maximo'],
                'unidad': row['unidad'],
                'establecido': row['establecido'],
                'categoria_regulacion': row['categoria_regulacion'],
                'referencias': row['referencias_texto']  # Usar el texto completo
            })
        
        # Información de paginación
        total_pages = (total_records + page_size - 1) // page_size if total_records > 0 else 1
        
        return {
            "success": True,
            "table_data": table_data,
            "pagination": {
                "current_page": page,
                "total_pages": total_pages,
                "page_size": page_size,
                "has_previous": page > 1,
                "has_next": page < total_pages,
                "showing_from": start_idx + 1 if total_records > 0 else 0,
                "showing_to": min(end_idx, total_records),
                "total_records": total_records
            }
        }
    except Exception as e:
        print(f"Error en suplementos-analysis: {e}")
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "error": str(e),
            "table_data": [],
            "pagination": {
                "current_page": 1,
                "total_pages": 1,
                "page_size": page_size,
                "has_previous": False,
                "has_next": False,
                "showing_from": 0,
                "showing_to": 0,
                "total_records": 0
            }
        }

@app.get("/api/suplementos-comparison")
async def get_suplementos_comparison(
    request: Request,
    paises: str,
    categorias: str,
    user: str = Depends(require_auth)
):
    """API para comparación regulatoria entre países"""
    try:
        pais_list = [p.strip() for p in paises.split(',') if p.strip()]
        categoria_list = [c.strip() for c in categorias.split(',') if c.strip()]
        
        print(f"🔍 Comparación solicitada:")
        print(f"   - Países: {pais_list}")
        print(f"   - Categorías: {categoria_list}")
        
        # Obtener datos regulatorios del archivo regulatory_data.py
        datos_regulatorios = extraer_info_regulatoria_pdf()
        print(f"Datos regulatorios disponibles para países: {list(datos_regulatorios.keys())}")
        
        comparison_data = {}
        
        for pais in pais_list:
            comparison_data[pais] = {}
            
            # Verificar si el país existe en los datos regulatorios
            if pais not in datos_regulatorios:
                print(f"⚠️ País '{pais}' no encontrado en datos regulatorios")
                comparison_data[pais] = {categoria: {} for categoria in categoria_list}
                continue
            
            for categoria in categoria_list:
                print(f"   📋 Procesando categoría: {categoria}")
                
                # Obtener subcategorías de la categoría seleccionada
                subcategorias = get_subcategories(categoria)
                print(f"      - Subcategorías encontradas: {list(subcategorias.keys())}")
                
                comparison_data[pais][categoria] = {}
                
                for subcategoria_nombre, subcategoria_key in subcategorias.items():
                    # Obtener información del país para esta subcategoría
                    info = get_regulatory_info(pais, subcategoria_key)
                    comparison_data[pais][categoria][subcategoria_nombre] = info
                    print(f"         {subcategoria_nombre}: {len(str(info))} caracteres")
        
        print(f"✅ Comparación completada para {len(pais_list)} países y {len(categoria_list)} categorías")
        
        return {
            "success": True,
            "comparison_data": comparison_data,
            "available_countries": get_available_countries(),
            "available_categories": get_available_categories()
        }
        
    except Exception as e:
        print(f"❌ Error en suplementos-comparison: {e}")
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "error": str(e),
            "comparison_data": {},
            "available_countries": [],
            "available_categories": []
        }


@app.get("/api/suplementos-categories")
async def get_suplementos_categories(request: Request, user: str = Depends(require_auth)):
    """API para obtener las categorías regulatorias disponibles - CORREGIDA"""
    try:
        print("Obteniendo categorías regulatorias desde regulatory_data.py")
        
        # Usar las funciones del archivo regulatory_data.py
        categories = get_available_categories()
        countries = get_available_countries()
        
        print(f"Categorías encontradas: {categories}")
        print(f"Países encontrados: {countries}")
        # Limpiar cualquier categoría incorrecta que pueda estar en cache
        print(f"Estructura de categorías: {CATEGORIAS_REGULATORIAS}")
        return {
            "success": True,
            "categories": categories,
            "countries": countries,
            "categories_structure": CATEGORIAS_REGULATORIAS
        }
    except Exception as e:
        print(f"Error en suplementos-categories: {e}")
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "error": str(e),
            "categories": [],
            "countries": [],
            "categories_structure": {}
        }


# AGREGAR API PARA ANÁLISIS DE OTRAS SUSTANCIAS
@app.get("/api/otras-sustancias-analysis")
async def get_otras_sustancias_analysis(
    request: Request,
    categoria: Optional[str] = None,
    paises: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    user: str = Depends(require_auth)
):
    """API para análisis de otras sustancias con filtros y paginación"""
    try:
        global df_otras_sustancias_cache
        
        if df_otras_sustancias_cache is None:
            df_otras_sustancias_cache = create_sample_otras_sustancias_data()
        
        # Crear una copia para trabajar
        filtered_df = df_otras_sustancias_cache.copy()
        
        # Filtrar por categoría
        if categoria and categoria != "all":
            filtered_df = filtered_df[filtered_df['Category'] == categoria]
        
        # Transformar datos de formato wide a long para facilitar filtros por país
        country_columns = [col for col in filtered_df.columns 
                          if col not in ['Category', 'Substance']]
        
        # Filtrar por países si se especifica
        if paises and paises.strip():
            pais_list = [p.strip() for p in paises.split(',') if p.strip()]
            # Filtrar columnas de países
            valid_countries = [p for p in pais_list if p in country_columns]
            if valid_countries:
                # Mantener Category, Substance y los países seleccionados
                columns_to_keep = ['Category', 'Substance'] + valid_countries
                filtered_df = filtered_df[columns_to_keep]
            else:
                # Si no hay países válidos, mantener solo Category y Substance
                filtered_df = filtered_df[['Category', 'Substance']]
        
        # Convertir a formato long para la tabla
        if len(filtered_df) > 0:
            # Obtener columnas de países disponibles en los datos filtrados
            current_country_columns = [col for col in filtered_df.columns 
                                     if col not in ['Category', 'Substance']]
            
            # Transformar de wide a long
            table_data = []
            for _, row in filtered_df.iterrows():
                for country in current_country_columns:
                    table_data.append({
                        'categoria': row['Category'],
                        'sustancia': row['Substance'],
                        'pais': country,
                        'estatus': row[country]
                    })
            
            # Convertir de vuelta a DataFrame para paginación
            table_df = pd.DataFrame(table_data)
        else:
            table_df = pd.DataFrame(columns=['categoria', 'sustancia', 'pais', 'estatus'])
        
        # Paginación
        total_records = len(table_df)
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_df = table_df.iloc[start_idx:end_idx]
        
        # Convertir a dict para la respuesta
        table_data_response = paginated_df.to_dict('records') if total_records > 0 else []
        
        # Información de paginación
        total_pages = (total_records + page_size - 1) // page_size if total_records > 0 else 1
        
        return {
            "success": True,
            "table_data": table_data_response,
            "pagination": {
                "current_page": page,
                "total_pages": total_pages,
                "page_size": page_size,
                "has_previous": page > 1,
                "has_next": page < total_pages,
                "showing_from": start_idx + 1 if total_records > 0 else 0,
                "showing_to": min(end_idx, total_records),
                "total_records": total_records
            }
        }
        
    except Exception as e:
        print(f"Error en otras-sustancias-analysis: {e}")
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "error": str(e),
            "table_data": [],
            "pagination": {
                "current_page": 1,
                "total_pages": 1,
                "page_size": page_size,
                "has_previous": False,
                "has_next": False,
                "showing_from": 0,
                "showing_to": 0,
                "total_records": 0
            }
        }


@app.get("/api/suplementos-export-analysis")
async def export_suplementos_analysis(
    request: Request,
    tipo: Optional[str] = None,
    ingredientes: Optional[str] = None,
    paises: Optional[str] = None,
    user: str = Depends(require_auth)
):
    """Exportar datos de análisis de suplementos a CSV"""
    global df_suplementos_cache, df_referencias_cache
    
    # Aplicar mismos filtros que en el análisis
    filtered_df = df_suplementos_cache.copy()
    
    if tipo and tipo != "all":
        filtered_df = filtered_df[filtered_df['tipo'] == tipo]
    
    if ingredientes and ingredientes.strip():
        ingrediente_list = [i.strip() for i in ingredientes.split(',') if i.strip()]
        if ingrediente_list:
            filtered_df = filtered_df[filtered_df['ingrediente'].isin(ingrediente_list)]
    
    if paises and paises.strip():
        pais_list = [p.strip() for p in paises.split(',') if p.strip()]
        if pais_list:
            filtered_df = filtered_df[filtered_df['pais'].isin(pais_list)]
    
    # APLICAR MISMO JOIN QUE EN LA FUNCIÓN DE ANÁLISIS
    def get_reference_text(ref_num, tipo):
        if pd.isna(ref_num) or ref_num == 0:
            return "-"
        
        try:
            ref_num = int(ref_num)
            ref_row = df_referencias_cache[
                (df_referencias_cache['referencia'] == ref_num) & 
                (df_referencias_cache['tipo'] == tipo)
            ]
            
            if not ref_row.empty:
                descripcion = ref_row.iloc[0]['descripcion']
                return f"{ref_num}: {descripcion}"
            else:
                return f"{ref_num}: Referencia no encontrada"
        except:
            return str(ref_num)
    
    # Aplicar el JOIN
    filtered_df = filtered_df.copy()
    filtered_df['referencias'] = filtered_df.apply(
        lambda row: get_reference_text(row['referencias'], row['tipo']), 
        axis=1
    )

    # NUEVA CORRECCIÓN: Convertir True/False a Sí/No
    filtered_df['establecido'] = filtered_df['establecido'].map({
        True: 'Sí',
        False: 'No'
    })
    
    # Generar Excel con formato
    excel_output = create_branded_excel(
        df=filtered_df,
        title="Tablero de Suplementos Alimenticios - América Latina",
        subtitle='Análisis de regulación de suplementos "alimenticios" por país',
        section_name="Análisis de Suplementos",
    )

    filename = f"suplementos_analisis_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        excel_output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Descarga de datos de comparación regulatoria

@app.get("/api/suplementos-export-comparison")
async def export_suplementos_comparison(
    request: Request,
    paises: str,
    categorias: str,
    user: str = Depends(require_auth)
):
    """Exportar datos de comparación regulatoria a CSV"""
    try:
        pais_list = [p.strip() for p in paises.split(',') if p.strip()]
        categoria_list = [c.strip() for c in categorias.split(',') if c.strip()]
        
        print(f"Exportando comparación para países: {pais_list}")
        print(f"Exportando comparación para categorías: {categoria_list}")
        
        # CORRECCIÓN: Usar los mismos datos regulatorios que en get_suplementos_comparison()
        from regulatory_data import extraer_info_regulatoria_pdf, get_subcategories, get_regulatory_info
        
        # Obtener datos regulatorios del archivo regulatory_data.py
        datos_regulatorios = extraer_info_regulatoria_pdf()
        
        # Crear lista para almacenar los datos en formato tabular
        export_data = []
        
        for pais in pais_list:
            # Verificar si el país existe en los datos regulatorios
            if pais not in datos_regulatorios:
                print(f"País '{pais}' no encontrado en datos regulatorios")
                # Agregar fila vacía para países sin datos
                for categoria in categoria_list:
                    subcategorias = get_subcategories(categoria)
                    for subcategoria_nombre, subcategoria_key in subcategorias.items():
                        export_data.append({
                            'pais': pais,
                            'categoria': categoria,
                            'subcategoria': subcategoria_nombre,
                            'informacion_regulatoria': 'Información no disponible para este país'
                        })
                continue
            
            for categoria in categoria_list:
                print(f"Procesando categoría: {categoria}")
                
                # Obtener subcategorías de la categoría seleccionada
                subcategorias = get_subcategories(categoria)
                print(f"Subcategorías encontradas: {list(subcategorias.keys())}")
                
                for subcategoria_nombre, subcategoria_key in subcategorias.items():
                    # Obtener información del país para esta subcategoría
                    info = get_regulatory_info(pais, subcategoria_key)
                    
                # Limpiar el texto de información regulatoria para CSV
                # Remover markdown y formatear para texto plano
                info_clean = info

                # Si info es un array (como las frases de advertencia), convertir a string
                if isinstance(info, list):
                    info_clean = '; '.join(str(item) for item in info)
                elif info_clean and info_clean != 'Información no disponible':
                    # Convertir a string si no lo es
                    info_clean = str(info_clean)
                    
                    # Remover enlaces markdown [texto](url)
                    import re
                    info_clean = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', info_clean)
                    # Remover texto en negrita **texto**
                    info_clean = re.sub(r'\*\*(.*?)\*\*', r'\1', info_clean)
                    # Reemplazar saltos de línea con punto y coma para CSV
                    info_clean = info_clean.replace('\n', '; ')
                    # Limpiar múltiples espacios
                    info_clean = re.sub(r'\s+', ' ', info_clean).strip()
                    
                    export_data.append({
                        'pais': pais,
                        'categoria': categoria,
                        'subcategoria': subcategoria_nombre,
                        'informacion_regulatoria': info_clean or 'Información no disponible'
                    })
                    
                    print(f"   {subcategoria_nombre}: {len(str(info_clean))} caracteres")
        
        # Convertir a DataFrame
        import pandas as pd
        df_export = pd.DataFrame(export_data)
        
        # Ordenar por país y categoría
        df_export = df_export.sort_values(['pais', 'categoria', 'subcategoria'])
        
        print(f"Datos preparados para exportación: {len(df_export)} filas")
        
        # Generar Excel con formato
        excel_output = create_branded_excel(
            df=df_export,
            title="Tablero de Suplementos Alimenticios - América Latina",
            subtitle='Análisis de regulación de suplementos "alimenticios" por país',
            section_name="Comparación Regulatoria",
            column_widths={4: 80}
        )

        filename = f"suplementos_comparacion_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        return StreamingResponse(
            excel_output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"Error en export_suplementos_comparison: {e}")
        import traceback
        traceback.print_exc()

        error_data = pd.DataFrame([{
            'Error': 'Error al generar exportación',
            'Detalle': str(e)
        }])

        excel_output = create_branded_excel(
            df=error_data,
            title="Error en Exportación",
            subtitle="Se produjo un error al generar el archivo",
            section_name="Error",
        )

        return StreamingResponse(
            excel_output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=suplementos_comparacion_error.xlsx"}
        )

# AGREGAR API PARA EXPORTAR OTRAS SUSTANCIAS
@app.get("/api/otras-sustancias-export")
async def export_otras_sustancias(
    request: Request,
    categoria: Optional[str] = None,
    paises: Optional[str] = None,
    user: str = Depends(require_auth)
):
    """Exportar datos de otras sustancias a CSV"""
    try:
        global df_otras_sustancias_cache
        
        # Aplicar mismos filtros que en el análisis
        filtered_df = df_otras_sustancias_cache.copy()
        
        if categoria and categoria != "all":
            filtered_df = filtered_df[filtered_df['Category'] == categoria]
        
        # Filtrar por países
        country_columns = [col for col in filtered_df.columns 
                          if col not in ['Category', 'Substance']]
        
        if paises and paises.strip():
            pais_list = [p.strip() for p in paises.split(',') if p.strip()]
            valid_countries = [p for p in pais_list if p in country_columns]
            if valid_countries:
                columns_to_keep = ['Category', 'Substance'] + valid_countries
                filtered_df = filtered_df[columns_to_keep]
        
        # Generar Excel con formato
        excel_output = create_branded_excel(
            df=filtered_df,
            title="Tablero de Suplementos Alimenticios - América Latina",
            subtitle="Otras sustancias reguladas por país",
            section_name="Otras Sustancias",
        )

        filename = f"otras_sustancias_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        return StreamingResponse(
            excel_output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"Error en export_otras_sustancias: {e}")

        error_data = pd.DataFrame([{
            'Error': 'Error al generar exportación',
            'Detalle': str(e)
        }])

        excel_output = create_branded_excel(
            df=error_data,
            title="Error en Exportación",
            subtitle="Se produjo un error al generar el archivo",
            section_name="Error",
        )

        return StreamingResponse(
            excel_output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=otras_sustancias_error.xlsx"}
        )

# AGREGAR AL FINAL DE main.py, ANTES DE LAS APIs DE OTRAS SUSTANCIAS

def load_actualizacion_regulatoria_from_excel():
    """Carga datos reales de actualización regulatoria desde el archivo Excel"""
    try:
        import pandas as pd
        from datetime import datetime
        
        # Leer el archivo Excel
        df = pd.read_excel('docs_en_act_reg_cons_pub.xlsx', sheet_name='Hoja1')
        
        print(f"Datos de actualización regulatoria cargados: {len(df)} registros")
        print(f"Columnas: {list(df.columns)}")
        
        # Limpiar y procesar los datos
        data = []
        
        for _, row in df.iterrows():
            # Limpiar espacios en blanco
            pais = str(row['PAÍS']).strip() if pd.notna(row['PAÍS']) else ''
            alcance = str(row['ALCANCE - BORRADOR']).strip() if pd.notna(row['ALCANCE - BORRADOR']) else ''
            normatividad = str(row['NORMATIVIDAD EN REVISIÓN']).strip() if pd.notna(row['NORMATIVIDAD EN REVISIÓN']) else ''
            link_docs = str(row['LINK / DOCUMENTOS']).strip() if pd.notna(row['LINK / DOCUMENTOS']) else ''
            
            # Procesar fecha
            fecha_str = ''
            if pd.notna(row['FECHA']):
                try:
                    if isinstance(row['FECHA'], datetime):
                        fecha_str = row['FECHA'].strftime('%Y-%m-%d')
                    else:
                        # Intentar parsear la fecha si es string
                        fecha_obj = pd.to_datetime(row['FECHA'])
                        fecha_str = fecha_obj.strftime('%Y-%m-%d')
                except:
                    fecha_str = str(row['FECHA'])
            
            # Solo agregar si tiene datos válidos
            if pais and alcance:
                data.append({
                    'pais': pais,
                    'alcance_borrador': alcance,
                    'normatividad_revision': normatividad,
                    'fecha': fecha_str,
                    'link_documentos': link_docs
                })
        
        print(f"Registros procesados exitosamente: {len(data)}")
        return data
        
    except FileNotFoundError:
        print("⚠️ Archivo docs_en_act_reg_cons_pub.xlsx no encontrado. Usando datos de ejemplo.")
        return create_sample_actualizacion_regulatoria_data()
    except Exception as e:
        print(f"Error cargando datos de actualización regulatoria: {e}")
        print("Usando datos de ejemplo...")
        return create_sample_actualizacion_regulatoria_data()

def create_sample_actualizacion_regulatoria_data():
    """Función de respaldo con datos de ejemplo (mantener por si falla la carga del Excel)"""
    import random
    from datetime import datetime, timedelta
    
    paises = ['Argentina', 'Brasil', 'Chile', 'Colombia', 'México', 'Perú', 'Costa Rica', 'Ecuador']
    
    alcances = [
        'Límites máximos de vitaminas y minerales',
        'Lista de ingredientes permitidos',
        'Declaraciones de propiedades saludables',
        'Etiquetado nutricional',
        'Requisitos de registro sanitario'
    ]
    
    normatividades = [
        'Resolución sobre límites de micronutrientes',
        'Normativa técnica de suplementos dietarios',
        'Reglamento de productos naturales',
        'Directrices de etiquetado'
    ]
    
    documentos = [
        'Consulta pública - Borrador disponible',
        'Proyecto de norma en revisión',
        'Documento técnico preliminar'
    ]
    
    data = []
    
    for pais in paises[:4]:  # Solo 4 países de ejemplo
        fecha_base = datetime.now()
        dias_adelante = random.randint(0, 180)
        fecha = fecha_base + timedelta(days=dias_adelante)
        
        data.append({
            'pais': pais,
            'alcance_borrador': random.choice(alcances),
            'normatividad_revision': random.choice(normatividades),
            'fecha': fecha.strftime('%Y-%m-%d'),
            'link_documentos': random.choice(documentos)
        })
    
    return data

# API para datos iniciales de actualización regulatoria
@app.get("/api/actualizacion-regulatoria-initial")
async def get_actualizacion_regulatoria_initial(request: Request, user: str = Depends(require_auth)):
    """API para obtener datos iniciales de actualización regulatoria"""
    try:
        global df_actualizacion_regulatoria_cache
        
        # Usar cache si está disponible, sino cargar desde archivo
        if df_actualizacion_regulatoria_cache is None:
            df_actualizacion_regulatoria_cache = load_actualizacion_regulatoria_from_excel()
        
        data = df_actualizacion_regulatoria_cache
        
        # Obtener lista de países únicos
        paises = sorted(list(set(item['pais'] for item in data if item['pais'])))
        
        return {
            "success": True,
            "data": data,
            "countries": paises,
            "total_records": len(data)
        }
        
    except Exception as e:
        print(f"Error en actualizacion-regulatoria-initial: {e}")
        return {
            "success": False,
            "error": str(e),
            "data": [],
            "countries": [],
            "total_records": 0
        }

# API para análisis de actualización regulatoria con filtros
@app.get("/api/actualizacion-regulatoria-analysis")
async def get_actualizacion_regulatoria_analysis(
    request: Request,
    paises: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    user: str = Depends(require_auth)
):
    """API para análisis de actualización regulatoria con filtros y paginación"""
    try:
        global df_actualizacion_regulatoria_cache
        
        # Usar cache si está disponible, sino cargar desde archivo
        if df_actualizacion_regulatoria_cache is None:
            df_actualizacion_regulatoria_cache = load_actualizacion_regulatoria_from_excel()
        
        all_data = df_actualizacion_regulatoria_cache.copy()
        
        # Aplicar filtros
        filtered_data = all_data.copy()
        
        if paises and paises.strip():
            pais_list = [p.strip() for p in paises.split(',') if p.strip()]
            if pais_list:
                filtered_data = [item for item in filtered_data if item['pais'] in pais_list]
        
        # Ordenar por fecha (más recientes primero)
        filtered_data.sort(key=lambda x: x['fecha'] if x['fecha'] else '1900-01-01', reverse=True)
        
        # Paginación
        total_records = len(filtered_data)
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        paginated_data = filtered_data[start_idx:end_idx]
        
        # Información de paginación
        total_pages = (total_records + page_size - 1) // page_size if total_records > 0 else 1
        
        return {
            "success": True,
            "table_data": paginated_data,
            "pagination": {
                "current_page": page,
                "total_pages": total_pages,
                "page_size": page_size,
                "has_previous": page > 1,
                "has_next": page < total_pages,
                "showing_from": start_idx + 1 if total_records > 0 else 0,
                "showing_to": min(end_idx, total_records),
                "total_records": total_records
            }
        }
        
    except Exception as e:
        print(f"Error en actualizacion-regulatoria-analysis: {e}")
        return {
            "success": False,
            "error": str(e),
            "table_data": [],
            "pagination": {
                "current_page": 1,
                "total_pages": 1,
                "page_size": page_size,
                "has_previous": False,
                "has_next": False,
                "showing_from": 0,
                "showing_to": 0,
                "total_records": 0
            }
        }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)